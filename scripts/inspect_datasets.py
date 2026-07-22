"""Inspect CSV/XLSX/DTA files without loading large datasets fully.

Outputs:
- docs/generated/dataset_catalog.csv
- docs/generated/variable_catalog.csv

Usage:
    python scripts/inspect_datasets.py --raw-dir data/raw --out-dir docs/generated

Notes:
- DTA inspection uses pandas' StataReader metadata reader.
- This script does not redistribute raw data; it only records metadata.
"""
from __future__ import annotations

import argparse
import subprocess
import warnings
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from pandas.io.stata import StataReader


def likely_owner(filename: str) -> str:
    name = filename.lower()
    if "boundary" in name:
        return "NSDI / RLMUA or administrative geospatial source"
    if any(token in name for token in ["nisr", "phc", "lfs", "vup", "population", "observationdata"]):
        return "NISR / Rwanda National Data Archive or official statistics source"
    if "movement" in name:
        return "Movement data provider / public mobility dataset"
    if "osm" in name or name.endswith(".pbf"):
        return "OpenStreetMap contributors"
    return "Unknown"


def recommended_layer(filename: str) -> str:
    name = filename.lower()
    if "boundary" in name:
        return "geo/admin_boundary"
    if "rwa_pd" in name:
        return "geo/population_density_grid"
    if "population_count" in name or "phc" in name:
        return "curated/population_demographic"
    if "lfs" in name:
        return "curated/labour_economic"
    if "vup" in name:
        return "curated/household_welfare"
    if "movement" in name:
        return "curated/mobility"
    if "ec-2023" in name or "establishment" in name:
        return "curated/business_environment"
    if "osm" in name or name.endswith(".pbf"):
        return "geo/osm"
    return "raw/statistical_indicator"


def relevance(filename: str) -> str:
    name = filename.lower()
    if "ec-2023" in name or "establishment" in name:
        return "High: business density, category distribution, commercial activity and formal/informal enterprise patterns."
    if "phc" in name:
        return "High: demographics, households, housing, education and employment proxies after aggregation."
    if "lfs" in name:
        return "Medium-High: employment, income and labour-market context by district/province."
    if "vup_s5f" in name:
        return "High: access-to-services distance/time indicators."
    if "expenditure" in name or "vup_s8" in name:
        return "Medium: expenditure and purchasing-power proxy after aggregation."
    if "vup_s9" in name:
        return "Low-Medium: welfare/income-support context; use aggregated indicators only."
    if "population_count" in name:
        return "High: sector-level population, growth and gender share features."
    if "rwa_pd" in name:
        return "Very high: fine-scale population density around candidate locations."
    if "boundary" in name:
        return "High for spatial joins if geometry is available; CSV attributes alone do not enable polygon joins."
    if "movement" in name:
        return "Medium: district-level mobility context; not street-level foot traffic."
    return "Inspect before use."


def count_csv_rows(path: Path) -> int | None:
    try:
        output = subprocess.check_output(["bash", "-lc", f"wc -l < {str(path)!r}"]).decode().strip()
        return max(0, int(output) - 1)
    except Exception:
        return None


def inspect_csv(path: Path) -> tuple[int | None, list[str], list[dict[str, Any]], dict[str, str]]:
    df = pd.read_csv(path, nrows=5, low_memory=False)
    sample = df.head(2).astype(str).replace({"nan": ""}).to_dict(orient="records")
    return count_csv_rows(path), list(df.columns), sample, {}


def inspect_xlsx(path: Path) -> tuple[int | None, list[str], list[dict[str, Any]], dict[str, str]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(min_row=1, max_row=5, values_only=True))
    columns = [str(c) if c is not None else "" for c in rows[0]] if rows else []
    sample = [dict(zip(columns, [str(v) if v is not None else "" for v in row])) for row in rows[1:3]]
    return ws.max_row - 1, columns, sample, {}


def inspect_dta(path: Path) -> tuple[int | None, list[str], list[dict[str, Any]], dict[str, str]]:
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        reader = StataReader(str(path))
        labels = reader.variable_labels()
        df = reader.read(nrows=2)
    columns = list(labels.keys()) or list(df.columns)
    sample = df.head(2).astype(str).replace({"nan": ""}).to_dict(orient="records")
    return getattr(reader, "_nobs", None), columns, sample, labels


def inspect_file(path: Path) -> tuple[int | None, list[str], list[dict[str, Any]], dict[str, str], str]:
    suffix = path.suffix.lower()
    try:
        if suffix == ".csv":
            rows, cols, sample, labels = inspect_csv(path)
        elif suffix == ".xlsx":
            rows, cols, sample, labels = inspect_xlsx(path)
        elif suffix == ".dta":
            rows, cols, sample, labels = inspect_dta(path)
        else:
            rows, cols, sample, labels = None, [], [], {}
        return rows, cols, sample, labels, ""
    except Exception as exc:
        return None, [], [], {}, f"Inspection warning: {type(exc).__name__}: {exc}"


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--raw-dir", default="data/raw")
    parser.add_argument("--out-dir", default="docs/generated")
    args = parser.parse_args()

    raw_dir = Path(args.raw_dir)
    out_dir = Path(args.out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    dataset_rows: list[dict[str, Any]] = []
    variable_rows: list[dict[str, Any]] = []

    for path in sorted(raw_dir.glob("**/*")):
        if path.is_dir() or path.suffix.lower() not in {".csv", ".dta", ".xlsx", ".pbf", ".geojson", ".gpkg", ".shp"}:
            continue
        rows, columns, _sample, labels, note = inspect_file(path)
        key = path.stem.lower().replace(" ", "_").replace("-", "_")[:120]
        size_mb = path.stat().st_size / (1024 * 1024)

        dataset_rows.append(
            {
                "dataset_key": key,
                "filename": path.name,
                "format": path.suffix.lower().replace(".", "").upper(),
                "size_mb": round(size_mb, 2),
                "rows_estimate": rows,
                "columns_count": len(columns),
                "owner_likely": likely_owner(path.name),
                "license_status": "verify_before_use",
                "permission_status": "not_confirmed",
                "recommended_layer": recommended_layer(path.name),
                "relevance": relevance(path.name),
                "sample_columns": ", ".join(columns[:15]),
                "inspection_note": note,
            }
        )

        for i, col in enumerate(columns):
            variable_rows.append(
                {
                    "dataset_key": key,
                    "filename": path.name,
                    "position": i + 1,
                    "variable_name": col,
                    "variable_label": labels.get(col, ""),
                    "recommended_use": "inspect",
                    "feature_candidate": "",
                    "notes": "",
                }
            )

    pd.DataFrame(dataset_rows).to_csv(out_dir / "dataset_catalog.csv", index=False)
    pd.DataFrame(variable_rows).to_csv(out_dir / "variable_catalog.csv", index=False)
    print(f"Wrote {len(dataset_rows)} datasets and {len(variable_rows)} variables to {out_dir}")


if __name__ == "__main__":
    main()
